"""ESD #2 fiscal model — web backend. The canonical Excel is the calculation engine:
writes inputs into a copy, recalculates with LibreOffice, reads Results back —
so the web numbers match the workbook exactly."""
import openpyxl, subprocess, tempfile, shutil, os, glob

TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "engine.xlsx")
SOFFICE  = os.environ.get("SOFFICE", "soffice")

ADV_SCALAR = {  # deep assumptions -> Advanced Inputs!D
 "city_pt_rate":"D10","esd_pt_rate":"D11","pop":"D13","pop_cagr":"D14","pool_base":"D15",
 "pool_cagr":"D16","av_cagr":"D17","capture":"D18","opex_infl":"D19","cap_infl":"D20",
 "city_res":"D21","esd_res":"D22","reserve_incl_cap":"D23","collection":"D24",
 "sfr_occ":"D26","mfr_occ":"D27","sfr_hh":"D28","mfr_hh":"D29","online_spend":"D30",
 "online_taxable":"D31","online_capture":"D32","online_growth":"D33","online_include":"D34",
 "far":"D47","abatement":"D48","abate_term":"D49","impact_fee":"D50"}
TYP_SPLIT  = {"city_share":"C5","edc":"C6","esd_share":"C7","pool_rate":"C8"}  # editable source
EXEMPT_ROW = {"Residential":38,"Commercial":39,"Industrial":40,"Civic":41,"Vacant":42}
RESULT_ROWS = {"pool":10,"city_salestax":11,"edc":12,"esd_salestax":13,"city_pt":18,"esd_pt":19,
 "impact":20,"city_opex":21,"city_capital":22,"esd_opex":23,"esd_capital":24,
 "city_net":25,"esd_net":26,"comb_net":27}
N_SCEN = 22  # Typical rows 13..34  /  Advanced rows 55..76

def _recalc(path):
    outdir = tempfile.mkdtemp()
    subprocess.run([SOFFICE,"--headless","--calc","--convert-to","xlsx","--outdir",outdir,path],
                   check=True, capture_output=True, timeout=180)
    shutil.copy(glob.glob(os.path.join(outdir,"*.xlsx"))[0], path); shutil.rmtree(outdir, ignore_errors=True)

def compute(inputs:dict, scenarios:list, exemptions:dict=None):
    tmp = tempfile.mktemp(suffix=".xlsx"); shutil.copy(TEMPLATE, tmp)
    wb = openpyxl.load_workbook(tmp); ADV = wb["Advanced Inputs"]; TYP = wb["Typical Inputs"]
    inputs = inputs or {}
    for k,cell in ADV_SCALAR.items():
        if inputs.get(k) is not None: ADV[cell] = inputs[k]
    if inputs.get("city_share") is not None:
        TYP["C5"] = inputs["city_share"]; TYP["C7"] = round(1-inputs["city_share"],6)
    if inputs.get("edc") is not None:       TYP["C6"] = inputs["edc"]
    if inputs.get("pool_rate") is not None: TYP["C8"] = inputs["pool_rate"]
    for cat,row in EXEMPT_ROW.items():
        e = (exemptions or {}).get(cat, {})
        for c in range(2,8):   ADV.cell(row,c).value = 0
        for c in range(10,16): ADV.cell(row,c).value = 0
        if "city" in e: ADV.cell(row,7).value  = e["city"]
        if "esd"  in e: ADV.cell(row,15).value = e["esd"]
    for i in range(N_SCEN):
        tr, ar = 13+i, 55+i
        if i < len(scenarios):
            s = scenarios[i]
            TYP.cell(tr,1).value = s.get("type","")
            TYP.cell(tr,2).value = "Yes" if s.get("active") else "No"
            TYP.cell(tr,3).value = s.get("qty")
            TYP.cell(tr,4).value = s.get("lot")
            TYP.cell(tr,5).value = s.get("grossup")
            ADV.cell(ar,6).value = s.get("manual_acres")
            ADV.cell(ar,7).value = s.get("start")
            ADV.cell(ar,8).value = s.get("buildout")
            ADV.cell(ar,9).value = s.get("capture")
        else:
            TYP.cell(tr,2).value = "No"
    wb.save(tmp); _recalc(tmp)
    rb = openpyxl.load_workbook(tmp, data_only=True); R = rb["Results"]; ci = rb["Capital Items"]
    out = {k:[R.cell(row,c).value for c in range(2,7)] for k,row in RESULT_ROWS.items()}
    capex=[]
    for r in range(4,33):
        if str(ci.cell(r,18).value).strip().lower()=="yes":
            capex.append({"item":ci.cell(r,1).value,"qty":ci.cell(r,12).value,"unit":ci.cell(r,4).value,
                          "payer":ci.cell(r,17).value,"treatment":ci.cell(r,19).value,
                          "cost10y":(ci.cell(r,25).value or 0)+(ci.cell(r,30).value or 0)})
    out["__capex__"]=capex; os.remove(tmp); return out

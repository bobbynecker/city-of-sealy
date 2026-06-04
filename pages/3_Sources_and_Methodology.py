"""
Sources & Methodology — where every number comes from.
Transparency page: data sources, design choices, and the live Sources & Assumptions
register read straight out of the fiscal-model workbook itself.
"""
import os

import streamlit as st
import _brand
import pandas as pd

st.set_page_config(page_title="Sealy — Sources & Methodology", page_icon="assets/favicon.png", layout="wide")
_brand.hide_chrome()
_brand.top_nav()

_brand.page_header(
    "Sources & methodology",
    "Every input in these tools is public, documented, and adjustable. This page shows where the "
    "numbers come from and the design choices behind the model — so you can check the math.")

st.subheader("How the engine works")
st.markdown(
    "1. **You set the inputs** — the City / ESD split, growth rates, exemptions, and the development scenario.\n"
    "2. **The app writes them into a copy of the actual fiscal-model workbook** (the same Excel file used "
    "for the Council analysis).\n"
    "3. **The workbook recalculates on the server** (LibreOffice, headless).\n"
    "4. **The app reads the Results block back.** Nothing is re-implemented in web code — the web results "
    "match the spreadsheet exactly, to the dollar."
)

st.subheader("Primary data sources")
st.dataframe(pd.DataFrame([
    ["Sales-tax sourcing rules", "Texas Comptroller Pub. 94-105 (Local Sales & Use Tax)", "Retail = origin; residential = online use-tax only; warehouse/manufacturing ≈ $0 local sales tax"],
    ["Sales-tax base & history", "Texas Comptroller allocation history", "City + ESD allocation trends; pool calibration"],
    ["Property values & exemptions", "Austin CAD", "City grants no general homestead; ESD #2 ≈ 1%; effective-haircut method"],
    ["City finances", "City of Sealy adopted budgets / ACFRs", "FY26 adopted rate $0.37731/$100 (Ord. 2025-24: M&O $0.20102 + I&S $0.17629)"],
    ["Population", "U.S. Census Bureau", "Sealy 7,937; growth CAGR adjustable"],
    ["ESD #2 finances", "ESD #2 budget (public record)", "Operating-cost calibration"],
    ["Per-acre coefficients", "Peer fiscal-impact literature", "22 development types, peer-calibrated and documented in the workbook"],
    ["Parcels & ownership", "Appraisal-district parcel data (Regrid compilation)", "5,463 parcels, Sealy city + ETJ, full geometry"],
    ["Land use & zoning", "City FLUM 2012–2035 + field-verified existing use", "Existing, adopted FLU, and rebalanced scenario layers"],
    ["Flood & pipelines", "FEMA NFHL · Texas RRC", "Live overlays, informational"],
], columns=["What", "Source", "How it's used"]), use_container_width=True, hide_index=True)

st.subheader("Design choices you should know about")
st.markdown(
    "- **The City / ESD split is blank by design.** It is the negotiation parameter — the model never "
    "assumes an answer.\n"
    "- **EDC retention defaults to 0%.** The Sealy EDC half-cent is shown as a separate carve-out, not "
    "buried in the City's share.\n"
    "- **Exemptions are an effective taxable-value haircut** per class (Residential / Commercial / "
    "Industrial), set against Austin CAD practice.\n"
    "- **Residential land is grossed up.** Infrastructure is sized off gross subdivision acres "
    "(lot size × gross-up ≈ 1.5× from measured Sealy subdivisions); home values and taxes stay per-home.\n"
    "- **The reserve line is operating-only** — simpler and defensible.\n"
    "- **Planning-level, not a certified forecast.** Inputs are documented and adjustable; this is not "
    "legal, financial, or official City advice."
)

st.subheader("The workbook's own sources register")
ENGINE = os.path.join(os.path.dirname(__file__), "..", "engine.xlsx")
try:
    from openpyxl import load_workbook
    wb = load_workbook(ENGINE, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "source" in s.lower()), None)
    if sheet:
        rows = [[("" if c is None else str(c)) for c in row]
                for row in wb[sheet].iter_rows(values_only=True)]
        rows = [row for row in rows if any(cell.strip() for cell in row)]
        if rows:
            width = max(len(row) for row in rows)
            rows = [row + [""] * (width - len(row)) for row in rows]
            with st.expander(f"View the live register from the workbook ('{sheet}' sheet)"):
                st.dataframe(pd.DataFrame(rows[1:], columns=[str(h) for h in rows[0]]),
                             use_container_width=True, hide_index=True)
    wb.close()
except Exception:
    st.caption("The live workbook register isn't available in this deployment — the summary above stands.")

st.caption("Model & parcel data last updated 2026-06-02. Questions or corrections are welcome — "
           "that's the point of publishing the math.")

_brand.footer()
